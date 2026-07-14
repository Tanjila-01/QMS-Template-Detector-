"""
Builds master_list.xlsx — a mock QMS Template Master List.

This is placeholder data (Alcon's real internal QMS list is proprietary and
not publicly accessible). The structure, ID format, and status logic mirror
what a real regulated-company template register looks like, so the code
built against it will work unchanged once the real Excel is provided.

Cross-check rules enforced before writing:
  1. No duplicate Template_ID values.
  2. Every Superseded_By value must itself exist as an "Effective" row.
  3. Withdrawn rows have no Superseded_By (a withdrawal means "stop using it",
     not "use this instead" — that's a business rule worth confirming with QA).
  4. Effective rows have no Superseded_By.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

rows = [
    # Template_ID,      Template_Name,                          Category,   Latest_Version, Status,      Effective_Date, Superseded_By,   Owner_Department
    ("V-QMS-0114221", "Deviation Management SOP",               "SOP",      3.0, "Effective",  "2025-08-01", "",              "Quality Assurance"),
    ("V-QMS-0114230", "Validation Master Plan Protocol",        "Protocol", 2.0, "Effective",  "2025-05-15", "",              "Validation Engineering"),
    ("V-QMS-0114245", "Design Verification Report",             "Report",   1.0, "Effective",  "2026-01-10", "",              "R&D Engineering"),
    ("V-QMS-0114260", "CAPA Initiation Form",                   "Form",     4.0, "Effective",  "2026-02-20", "",              "Quality Assurance"),
    ("V-QMS-0114275", "Risk Management Report (ISO 14971)",     "Report",   2.0, "Effective",  "2025-11-01", "",              "Risk Management"),
    ("V-QMS-0114100", "Deviation Management SOP (Legacy)",      "SOP",      2.0, "Superseded", "2022-03-01", "V-QMS-0114221", "Quality Assurance"),
    ("V-QMS-0114205", "Validation Master Plan Protocol (Old)",  "Protocol", 1.0, "Superseded", "2023-06-01", "V-QMS-0114230", "Validation Engineering"),
    ("V-QMS-0114050", "Change Control Form (Discontinued)",     "Form",     1.0, "Withdrawn",  "2021-01-01", "",              "Quality Assurance"),
    ("V-QMS-0114080", "Supplier Audit Checklist (Discontinued)","Form",     3.0, "Withdrawn",  "2020-09-15", "",              "Supplier Quality"),
    ("V-QMS-0114290", "Complaint Handling SOP",                 "SOP",      1.0, "Effective",  "2026-04-01", "",              "Post-Market Surveillance"),
]

# --- Cross-check before writing ---------------------------------------------
ids = [r[0] for r in rows]
assert len(ids) == len(set(ids)), "Duplicate Template_ID found!"

effective_ids = {r[0] for r in rows if r[4] == "Effective"}
for tid, name, cat, ver, status, eff_date, sup_by, dept in rows:
    if status == "Superseded":
        assert sup_by, f"{tid} is Superseded but has no Superseded_By"
        assert sup_by in effective_ids, f"{tid}'s replacement {sup_by} is not an Effective template"
    if status in ("Effective", "Withdrawn"):
        assert not sup_by, f"{tid} is {status} but has a Superseded_By set"

print(f"Cross-check passed: {len(rows)} rows, no duplicates, all Superseded_By references valid.")

# --- Build workbook -----------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Master List"

headers = ["Template_ID", "Template_Name", "Category", "Latest_Version",
           "Status", "Effective_Date", "Superseded_By", "Owner_Department"]
ws.append(headers)

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
body_font = Font(name="Arial", size=10)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

status_fill = {
    "Effective": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Superseded": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Withdrawn": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}
status_font = {
    "Effective": Font(name="Arial", size=10, color="006100"),
    "Superseded": Font(name="Arial", size=10, color="9C6500"),
    "Withdrawn": Font(name="Arial", size=10, color="9C0006"),
}

for c in ws[1]:
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

for row in rows:
    ws.append(row)

for r_idx in range(2, ws.max_row + 1):
    status_val = ws.cell(row=r_idx, column=5).value
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.font = body_font
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")
    status_cell = ws.cell(row=r_idx, column=5)
    status_cell.fill = status_fill[status_val]
    status_cell.font = status_font[status_val]
    status_cell.alignment = Alignment(horizontal="center", vertical="center")

widths = [16, 38, 10, 15, 12, 14, 16, 22]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws.freeze_panes = "A2"

# Legend sheet — documents assumptions per xlsx skill guidance
legend = wb.create_sheet("README")
legend_font = Font(name="Arial", size=10)
legend_title_font = Font(name="Arial", size=12, bold=True)
legend.append(["Master List — Notes & Assumptions"])
legend["A1"].font = legend_title_font
notes = [
    "",
    "This is a MOCK master list for prototype/demo purposes only.",
    "Alcon's real QMS template register is internal/proprietary — replace this file",
    "with the actual export from the QMS on presentation day; the app reads any",
    "Excel file with these exact column headers, so no code changes are needed.",
    "",
    "Column definitions:",
    "Template_ID       Unique QMS template identifier (format: V-QMS-XXXXXXX)",
    "Template_Name     Human-readable template title",
    "Category          SOP / Protocol / Report / Form",
    "Latest_Version    The current approved version number for this Template_ID",
    "Status            Effective / Superseded / Withdrawn",
    "Effective_Date    Date this status became active",
    "Superseded_By     If Superseded, the Template_ID that replaces it",
    "Owner_Department  QMS-registered owning department",
    "",
    "Business rules encoded here (confirm with QA before finalizing):",
    "- A document using an Effective Template_ID but an older version number",
    "  than Latest_Version => WARNING (newer version of same template exists).",
    "- A document using a Superseded or Withdrawn Template_ID => FAIL.",
    "- A Template_ID not found in this list at all => FLAG for manual QA check.",
]
for n in notes:
    legend.append([n])
for row in legend.iter_rows(min_row=2, max_row=legend.max_row):
    for cell in row:
        cell.font = legend_font
legend.column_dimensions["A"].width = 90

wb.save("/home/claude/template-detector/master_list.xlsx")
print("master_list.xlsx written.")
